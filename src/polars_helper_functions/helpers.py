"""Helper functions for working with Polars and Spyder."""

from __future__ import annotations


def save_schema(schema_or_frame, schema_path, infer_schema_length=None):
    """
    Save a Polars schema as JSON using dtype string representations.

    Parameters
    ----------
    schema_or_frame : Mapping[str, polars.DataType] | polars.Schema | polars.DataFrame | polars.LazyFrame
        Schema-like object to persist. This can be:
        - output of `collect_schema()`
        - a schema mapping
        - a `DataFrame`
        - a `LazyFrame`
    schema_path : str | pathlib.Path
        JSON destination path.
    infer_schema_length : int | None, default None
        Included for API compatibility with schema-inference workflows.
        This parameter is currently not used when the input is already a
        DataFrame/LazyFrame/schema object.
    """
    import json
    from pathlib import Path

    import polars as pl

    if isinstance(schema_or_frame, pl.LazyFrame):
        schema = schema_or_frame.collect_schema()
    elif isinstance(schema_or_frame, pl.DataFrame):
        schema = schema_or_frame.schema
    elif hasattr(schema_or_frame, "items"):
        schema = schema_or_frame
    else:
        raise TypeError(
            "`schema_or_frame` must be a Polars DataFrame/LazyFrame or a schema-like mapping."
        )

    _ = infer_schema_length  # reserved for compatibility with scan/read workflows

    schema_path = Path(schema_path)
    schema_path.parent.mkdir(parents=True, exist_ok=True)

    schema_to_save = {col: repr(dtype) for col, dtype in schema.items()}

    with schema_path.open("w", encoding="utf-8") as f:
        json.dump(schema_to_save, f, indent=2)


def load_saved_schema(schema_path):
    """
    Load a saved schema JSON and return a Polars-compatible schema mapping.

    Parameters
    ----------
    schema_path : str | pathlib.Path
        Path to JSON schema created by `save_schema`.

    Returns
    -------
    dict[str, polars.DataType]
        Schema dictionary that can be passed to `scan_csv(..., schema=...)`
        or `read_csv(..., schema=...)`.
    """
    import json
    from pathlib import Path

    import polars as pl

    schema_path = Path(schema_path)
    with schema_path.open(encoding="utf-8") as f:
        saved_schema = json.load(f)

    namespace = {
        name: getattr(pl, name)
        for name in (
            "Int8",
            "Int16",
            "Int32",
            "Int64",
            "UInt8",
            "UInt16",
            "UInt32",
            "UInt64",
            "Float32",
            "Float64",
            "String",
            "Boolean",
            "Date",
            "Time",
            "Datetime",
            "Duration",
            "Binary",
            "Decimal",
            "Categorical",
            "Enum",
            "Object",
            "Null",
            "Unknown",
            "List",
            "Array",
            "Struct",
        )
    }

    return {
        col: eval(dtype_repr, {"__builtins__": {}}, namespace)
        for col, dtype_repr in saved_schema.items()
    }


def show_unique(lf, cols, mode: str = "column"):
    """
    Show unique values from one or more columns in a Polars LazyFrame.

    Parameters
    ----------
    lf : polars.LazyFrame
        LazyFrame containing the columns.
    cols : str | list[str] | tuple[str, ...]
        One or more columns to inspect.
    mode : {"column", "row"}, default "column"
        - "column": shows unique values separately for each column
        - "row": shows unique row combinations across the selected columns
    """
    import polars as pl

    if not isinstance(cols, (list, tuple)):
        cols = [cols]

    if mode == "row":
        print(lf.select(cols).unique().collect())
    else:
        result = lf.select([pl.col(c).unique().sort().alias(c) for c in cols]).collect()
        print(result)


def view(df, n: int | None = 100, name: str = "a1_VIEW_DF"):
    """
    Send a Polars DataFrame/LazyFrame to Spyder Variable Explorer.

    Parameters
    ----------
    df : polars.DataFrame or polars.LazyFrame
    n : int or None, default 100
        Number of rows to materialize.
        If None, materialize all rows.
    name : str, default "a1_VIEW_DF"
        Variable name shown in Variable Explorer.
    """
    import __main__
    import polars as pl

    if isinstance(df, pl.LazyFrame):
        df = df.collect() if n is None else df.head(n).collect()
    elif n is not None:
        df = df.head(n)

    setattr(__main__, name, df)


def write_excel_polars(file_path, df, sheet_name, mode: str = "raw", table_name: str | None = None):
    """
    Write a Polars DataFrame to an Excel workbook using openpyxl.

    - If the workbook does not exist, create it
    - If the sheet exists, replace it
    - mode="raw": write plain cells
    - mode="table": write as an Excel table
    """
    from pathlib import Path

    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table

    file_path = Path(file_path)

    if file_path.exists():
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)
    ws.append(df.columns)

    for row in df.iter_rows():
        ws.append(row)

    if mode == "table":
        if table_name is None:
            table_name = f"tbl_{sheet_name.replace(' ', '_')}"

        nrows = df.height + 1
        ncols = df.width
        end_col = get_column_letter(ncols)
        table_ref = f"A1:{end_col}{nrows}"

        tab = Table(displayName=table_name, ref=table_ref)
        ws.add_table(tab)

    wb.save(file_path)


def clean_names(obj):
    """
    Clean column names into snake_case format.

    Parameters
    ----------
    obj : list[str], polars.DataFrame, or polars.LazyFrame
        Either:
        - a list of column names to clean, or
        - a Polars DataFrame/LazyFrame whose columns should be cleaned

    Returns
    -------
    list[str] or polars.DataFrame or polars.LazyFrame
        - If input is a list: returns cleaned list of column names
        - If input is a DataFrame/LazyFrame: returns object with renamed columns
    """
    import re

    import polars as pl

    def _clean(col):
        col = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", col)
        col = re.sub(r"([A-Z])([A-Z][a-z])", r"\1_\2", col)
        col = col.lower()
        col = re.sub(r"[\s\-\/\.]+", "_", col)
        col = re.sub(r"[^a-z0-9_]", "", col)
        col = re.sub(r"_+", "_", col)
        col = col.strip("_")
        return col

    if isinstance(obj, list):
        return [_clean(c) for c in obj]

    if isinstance(obj, (pl.DataFrame, pl.LazyFrame)):
        mapping = {c: _clean(c) for c in obj.columns}
        return obj.rename(mapping)

    raise TypeError("Input must be list[str], polars.DataFrame, or polars.LazyFrame.")


def clean_strings(df, cols=None, normalize: bool = False):
    """
    Clean string columns in a Polars DataFrame or LazyFrame.

    Default behavior:
    - strip leading/trailing whitespace
    - normalize unicode (NFKC)
    - remove control characters
    - collapse repeated whitespace

    Optional normalize=True:
    - convert to lowercase
    - remove all whitespace
    - remove punctuation
    - designed to make similar strings identical
    """
    import polars as pl
    import polars.selectors as cs

    if cols is None:
        cols = cs.string()

    expr = (
        pl.col(cols)
        .str.normalize("NFKC")
        .str.replace_all(r"[\x00-\x1F\x7F]", "")
        .str.replace_all(r"\s+", " ")
        .str.strip_chars()
    )

    if normalize:
        expr = expr.str.to_lowercase().str.replace_all(r"\s+", "").str.replace_all(r"[^\w]", "")

    return df.with_columns(expr)


def check_merge(left, right, on=None, left_on=None, right_on=None, view_unmatched: bool = False):
    """
    Print a Stata-like merge summary for two Polars DataFrames or LazyFrames.

    Notes
    -----
    This checks unique merge keys, not raw row counts.
    """
    import polars as pl

    if on is None and (left_on is None or right_on is None):
        raise ValueError("Provide either `on` or both `left_on` and `right_on`.")

    if on is not None:
        left_keys = [on] if isinstance(on, str) else list(on)
        right_keys = left_keys
        output_keys = left_keys
    else:
        left_keys = [left_on] if isinstance(left_on, str) else list(left_on)
        right_keys = [right_on] if isinstance(right_on, str) else list(right_on)
        output_keys = left_keys

    left = left.select(left_keys)
    right = right.select(right_keys)

    if hasattr(left, "collect"):
        left = left.collect()
    if hasattr(right, "collect"):
        right = right.collect()

    left_unique = left.unique()
    right_unique = right.unique()

    left_is_unique = left.height == left_unique.height
    right_is_unique = right.height == right_unique.height

    if left_is_unique and right_is_unique:
        merge_type = "1:1"
    elif not left_is_unique and right_is_unique:
        merge_type = "m:1"
    elif left_is_unique and not right_is_unique:
        merge_type = "1:m"
    else:
        merge_type = "m:m"

    if on is not None:
        matched = left_unique.join(right_unique, on=left_keys, how="inner").height
        left_only_df = left_unique.join(right_unique, on=left_keys, how="anti")
        right_only_df = right_unique.join(left_unique, on=right_keys, how="anti")
    else:
        matched = left_unique.join(
            right_unique,
            left_on=left_keys,
            right_on=right_keys,
            how="inner",
        ).height

        left_only_df = left_unique.join(
            right_unique,
            left_on=left_keys,
            right_on=right_keys,
            how="anti",
        )

        right_only_df = (
            right_unique.join(
                left_unique,
                left_on=right_keys,
                right_on=left_keys,
                how="anti",
            )
            .rename(dict(zip(right_keys, output_keys)))
            .select(output_keys)
        )

    left_only = left_only_df.height
    right_only = right_only_df.height
    total = matched + left_only + right_only

    print()
    print("Merge result".center(38))
    print("-" * 38)
    print(f"{'Not matched':<24}{left_only + right_only:>14,}")
    print(f"{'    from left only':<24}{left_only:>14,}")
    print(f"{'    from right only':<24}{right_only:>14,}")
    print(f"{'Matched':<24}{matched:>14,}")
    print("-" * 38)
    print(f"{'Total':<24}{total:>14,}")
    print("-" * 38)
    print(f"{'Merge type':<24}{merge_type:>14}")
    print()

    if view_unmatched:
        unmatched = pl.concat(
            [
                left_only_df.with_columns(pl.lit("left_only").alias("_merge")),
                right_only_df.with_columns(pl.lit("right_only").alias("_merge")),
            ],
            how="diagonal",
        )

        view(unmatched)
